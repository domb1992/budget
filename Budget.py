import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Liste der Monate, die verarbeitet werden sollen
months = ['Januar', 'Februar', 'Maerz', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember']

def extract_credit_card_data(pdf_file_path):
    transactions = []
    transaction_pattern = re.compile(r"(\d{2}\.\d{2}\.\d{2})\s+\d{2}\.\d{2}\.\d{2}\s+(.+?)\s+(\d{1,3}(?:\'\d{3})*\.\d{2})(?:\s+(\d{1,3}(?:\'\d{3})*\.\d{2}))?")
    special_transaction_pattern = re.compile(r"(\d{2}\.\d{2}\.\d{2})\s+(.+?)(?:IE|DE|US|CH)\s+CHF\s+(\d{1,3}(?:\'\d{3})*\.\d{2})\s+(\d{1,3}(?:\'\d{3})*\.\d{2})")
    conversion_pattern = re.compile(r"Umrechnungskurs.*?(\d{1,3}(?:\'\d{3})*\.\d{2})")
    fee_pattern = re.compile(r"Bearbeitungsgebühr.*?(\d{1,3}(?:\'\d{3})*\.\d{2})")

    with pdfplumber.open(pdf_file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                skip_next = False
                for i, line in enumerate(lines):
                    if skip_next:
                        skip_next = False
                        continue

                    special_match = special_transaction_pattern.search(line)
                    if special_match:
                        date = special_match.group(1)
                        description = special_match.group(2).strip()
                        amount_chf = float(special_match.group(4).replace("'", "").replace(",", ""))

                        transactions.append([date, description, amount_chf])
                        print(f"Special Transaction match: {date}, {description}, {amount_chf}")
                        continue

                    match = transaction_pattern.search(line)
                    if match:
                        date = match.group(1)
                        description = match.group(2).strip()
                        amount_original = float(match.group(3).replace("'", "").replace(",", ""))
                        amount_chf = float(match.group(4).replace("'", "").replace(",", "")) if match.group(4) else amount_original

                        if "Ihre Zahlung - Danke" not in description and "Totalbetrag letzte Abrechnung" not in description:
                            # Check for conversion and fee in the next lines
                            if i + 2 < len(lines):
                                next_line = lines[i + 1]
                                next_next_line = lines[i + 2]
                                if conversion_pattern.search(next_line) or fee_pattern.search(next_next_line):
                                    skip_next = True
                                    continue

                            transactions.append([date, description, amount_chf])
                            print(f"Transaction match: {date}, {description}, {amount_chf}")
                        continue

    return pd.DataFrame(transactions, columns=['Date', 'Description', 'Amount'])

# Liste zur Speicherung der zu verkettenden Datenrahmen initialisieren
frames = []

for month in months:
    # Dynamische Dateipfade erstellen
    csv_file_path = f'Pathto\\_Budget\\BANKFILE_{month}.xlsx'
    pdf_file_path = f'Pathto\\raiffeisen_export\\CREDITCARDFILE_{month}.pdf'
    
    # Verarbeiten der CSV-Datei
    try:
        bank_data = pd.read_excel(csv_file_path)
        bank_data = bank_data[['Booked At', 'Text', 'Credit/Debit Amount']]
        bank_data.columns = ['Date', 'Description', 'Amount']
        bank_data['Amount'] = bank_data['Amount'].astype(str).str.replace('-', '').astype(float).round(2)
        bank_data['Date'] = pd.to_datetime(bank_data['Date'], format='%Y-%m-%d')
        print(f"Verarbeitete Bankdaten für {month}:")
        print(bank_data)
        frames.append(bank_data)
    except FileNotFoundError:
        print(f"CSV-Datei für den Monat {month} nicht gefunden. Überspringen...")
    
    # Verarbeiten der PDF-Datei
    try:
        credit_card_data = extract_credit_card_data(pdf_file_path)
        credit_card_data['Date'] = pd.to_datetime(credit_card_data['Date'], format='%d.%m.%y')
        credit_card_data['Amount'] = credit_card_data['Amount'].astype(float).round(2)
        credit_card_data = credit_card_data.dropna(how='all')
        print(f"Verarbeitete Kreditkartendaten für {month}:")
        print(credit_card_data)
        frames.append(credit_card_data)
    except FileNotFoundError:
        print(f"PDF-Datei für den Monat {month} nicht gefunden. Überspringen...")

# Überprüfen, ob frames nicht leer ist
if frames:
    # Daten zusammenführen
    all_data = pd.concat(frames, ignore_index=True)
    
    # Pfad zur vorhandenen Datei und das Blatt "Ausgaben"
    existing_file_path = r'PathtoBUDGETFILE.xlsx'
    sheet_name = 'Ausgaben'
    
    # Lade die existierende Excel-Datei und die bestehenden Daten
    workbook = load_workbook(existing_file_path)
    sheet = workbook[sheet_name]
    
    # Bestehende Daten laden und in einen DataFrame umwandeln
    data = list(sheet.values)
    if data:
        cols = data[0]  # Erste Zeile als Spaltennamen verwenden
        existing_data = pd.DataFrame(data[1:], columns=cols)
    else:
        existing_data = pd.DataFrame(columns=['Date', 'Description', 'Amount', 'Category'])
    
    print("Bestehende Daten:")
    print(existing_data)
    
    # Zusammenführen der neuen und bestehenden Daten
    combined_data = pd.concat([existing_data, all_data], ignore_index=True).drop_duplicates(subset=['Date', 'Description', 'Amount'])
    
    print("Kombinierte Daten:")
    print(combined_data)
    
    # Zusätzliche Debugging-Information: Überprüfen der Zeile 291
    if combined_data.shape[0] > 290:
        print("Inhalt der Zeile 291 vor dem Export:")
        print(combined_data.iloc[290])
    
    # Lösche vorhandene Daten im Blatt "Ausgaben"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
    
    # Schreibe die kombinierten Daten in das Blatt "Ausgaben"
    for r_idx, row in combined_data.iterrows():
        for c_idx, value in enumerate(row):
            sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
            if r_idx == 290:  # Spezielle Debugging-Information für Zeile 291
                print(f"Writing to cell {r_idx + 2}, {c_idx + 1}: {value}")
    
    workbook.save(existing_file_path)
    
    print("Daten wurden erfolgreich verarbeitet und exportiert.")
else:
    print("Keine Daten zum Exportieren vorhanden.")

# Pause am Ende des Skripts
input("Drücken Sie die Eingabetaste, um das Skript zu beenden...")
